import openpyxl
from pathlib import Path

excel_path = Path('docs/excel/Comparison Table (SOA).xlsx')
wb = openpyxl.load_workbook(excel_path)
ws = wb[wb.sheetnames[0]]

dataset_map = {
    '1-s2.0-S0031320321003897-main (1)': 'SEN1-2 paired SAR→optical dataset; experiments use SAR VV polarization and RGB optical channels (4/3/2) with 256×256 patches.',
    '1505.04597v1': 'ISBI EM segmentation challenge data and ISBI 2015 cell tracking datasets (PhC-U373, DIC-HeLa), with strong augmentation.',
    '2309.07460v2 (1) (1)': 'Tutorial paper; no single fixed dataset. Discusses location-tagged channel/beam data sources for CKM in 6G systems.',
    '2402.00878v1 (2)': 'Open radio map estimation dataset (~74,515 maps, 424 areas): city geometry + Tx metadata + directional antennas + ray-tracing pathloss labels.',
    '2504.09849v1 (1) (1)': 'CKMImageNet: large-scale CKM dataset from ray tracing (urban maps, BS/Rx grids, path gain/AoA/AoD/delay labels) with 32/64/128 tiled variants.',
    '2507.12166v1': 'UrbanRadio3D: volumetric 3D×3D radio-map dataset (pathloss + DoA az/el + ToA), multi-height urban regions for diffusion benchmark.',
    '3.+18015+-+Aniru+Abudu+Nigeria+(311-332)': 'DeepMIMO, COST 2100, and NYU Wireless datasets used for training/evaluation of AI-based channel modeling.',
    'Electronics Letters - 2025 - Zhou - Vision Channel Knowledge Map‐Aide_00ae6c299c': 'ViWi/ViWi-cc style vision-wireless V2I dataset with synchronized camera and mmWave channel/beam data for VCKM beam selection.',
    'How_Much_Data_Is_Needed_for_Channel_Knowledge_Map_Construction': 'Analytical + simulation study on CKM/CGM sample density (random PPP and grid sampling); includes case study with sparse measured/simulated data.',
    'Isola_Image-To-Image_Translation_With_CVPR_2017_paper': 'Cityscapes, CMP Facades, map↔aerial imagery, edge→photo sets, and colorization data used across multiple image-translation tasks.',
    'project proposal and workplan.docx (6)': 'Project/workplan document; proposes future data from ray tracing + real measurements, not a finalized experimental dataset paper.',
    'sensors-24-00641': 'Wireless physical-layer authentication dataset from channel measurements transformed to CFR/CIR, augmented with GAN-generated synthetic samples.',
    'sensors-25-07463': 'Precision-agriculture IoT/UAV simulation environment data for MARL path planning (state trajectories, rewards, mission/task variables).',
    'SRC_SIg_21': 'Localization measurement dataset augmented with GAN-synthesized samples for DL-based localization model training and evaluation.',
    'Thesis Research_ Channel Knowledge Map Prediction': 'Thesis/proposal material; references planned CKM datasets from ray tracing and measured UAV-enabled 6G scenarios.',
    'UAV_height_distribution (1) (1)': 'Synthetic UAV altitude samples generated via bimodal Beta-mixture distribution (no external benchmark dataset).'
}

updated = 0
for r in range(2, ws.max_row + 1):
    key = ws.cell(r, 1).value
    if key in dataset_map:
        ws.cell(r, 10, dataset_map[key])
        updated += 1

wb.save(excel_path)
print('Dataset column updated rows:', updated)
print('Workbook:', excel_path)
