import openpyxl
from pathlib import Path

excel_path = Path('docs/excel/Comparison Table (SOA).xlsx')
wb = openpyxl.load_workbook(excel_path)
ws = wb[wb.sheetnames[0]]

# Manual LLM-authored enrichment for open fields
# col 9 Output, col 12 Heuristics, col 14 Disadvantages, col 15 Speed, col 16 Other
manual = {
    '1-s2.0-S0031320321003897-main (1)': {
        9: 'Comparative pattern-recognition results reported for the target task; performance is benchmarked against prior methods.',
        12: 'Focus on feature representation quality and model comparison protocol.',
        14: 'Generalization may depend on dataset/domain shift; requires careful feature/data preprocessing.',
        15: 'Not explicitly quantified in extracted text.',
        16: 'Useful as a methodological baseline paper.'
    },
    '1505.04597v1': {
        9: 'Strong segmentation quality in biomedical images; introduced U-Net as a highly effective encoder-decoder with skip connections.',
        12: 'Heavy use of data augmentation and patch-based training to improve robustness with limited labels.',
        14: 'Performance is data/domain dependent; may require task-specific tuning and substantial GPU resources for large 3D variants.',
        15: 'Inference is typically practical; training can be costly depending on image size and augmentation.',
        16: 'Foundational architecture with broad transfer value beyond medical segmentation.'
    },
    '2309.07460v2 (1) (1)': {
        9: 'Tutorial/survey paper: no single benchmark output, but a consolidated framework and open-problem analysis for CKM-enabled 6G.',
        12: 'System-level taxonomy: CKM construction + CKM utilization for reducing CSI overhead.',
        14: 'Main limitations are practical deployment and maintaining CKM under environment dynamics.',
        15: 'No direct runtime metric; focuses on conceptual and system-level efficiency.',
        16: 'High-value reference to position your own CKM pipeline and terminology.'
    },
    '2402.00878v1 (2)': {
        9: 'Provides radio-map dataset and initial estimation experiments; useful benchmark for directional-transmitter scenarios.',
        12: 'Data-centric benchmark construction plus baseline estimators for reproducibility.',
        14: 'Initial experiments may not cover all real-world deployment conditions.',
        15: 'Moderate computational demand for baseline models; scalable with model complexity.',
        16: 'Good bridge paper between raw data resources and model benchmarking.'
    },
    '2504.09849v1 (1) (1)': {
        9: 'Introduces CKMImageNet and demonstrates AI-oriented CKM prediction potential for comm+sensing tasks.',
        12: 'Dataset-first strategy with standardized splits and evaluation setup.',
        14: 'Dataset bias and environment coverage can limit transfer to unseen geographies.',
        15: 'Generation/training speed depends on model family; dataset itself enables faster iteration once preprocessed.',
        16: 'Key dataset paper for reproducible CKM experiments.'
    },
    '2507.12166v1': {
        9: 'Presents RadioDiff-3D and diffusion-based benchmark; demonstrates generative modeling for 3D radio-map completion/synthesis.',
        12: 'Generative diffusion pipeline with 3D×3D representation and benchmark protocol.',
        14: 'Diffusion models are compute-heavy and sensitive to training setup.',
        15: 'Training is relatively slow; inference can also be non-trivial vs. deterministic regressors.',
        16: 'Advanced but powerful direction for high-fidelity environment-aware modeling.'
    },
    '3.+18015+-+Aniru+Abudu+Nigeria+(311-332)': {
        9: 'Reports applied communication results with contextual analysis for the selected scenario.',
        12: 'Applied-method perspective with practical interpretation of observed metrics.',
        14: 'Likely limited by scope and dataset/measurement coverage.',
        15: 'Not clearly reported.',
        16: 'Useful contextual/region-focused supporting reference.'
    },
    'Electronics Letters - 2025 - Zhou - Vision Channel Knowledge Map‐Aide_00ae6c299c': {
        9: 'Shows CKM vision-aided beam selection gains in V2I vs. conventional alternatives.',
        12: 'Fuse visual/environment cues with beam-selection decision process.',
        14: 'Can be sensitive to camera quality, perception errors, and domain shift.',
        15: 'Intended for practical low-latency beam decisions, but exact runtime depends on vision stack.',
        16: 'Very aligned with your beam/CKM direction and applied relevance.'
    },
    'How_Much_Data_Is_Needed_for_Channel_Knowledge_Map_Construction': {
        9: 'Derives AMSE/data-density relationships and provides guidance for CKM sample requirements.',
        12: 'Model-based spatial prediction with analytical error derivation and local parameter estimation.',
        14: 'Assumptions on correlation/modeling can diverge from highly non-stationary real environments.',
        15: 'Main contribution is analytical guidance rather than runtime benchmarks.',
        16: 'Excellent paper for principled dataset sizing and CKM collection strategy.'
    },
    'Isola_Image-To-Image_Translation_With_CVPR_2017_paper': {
        9: 'Strong qualitative and quantitative gains for image-to-image tasks via conditional adversarial training (pix2pix).',
        12: 'cGAN objective + L1 loss + U-Net generator + PatchGAN discriminator.',
        14: 'Can hallucinate artifacts; training stability and mode behavior depend on setup/data.',
        15: 'Inference is generally fast once trained; training can be sensitive and time-consuming.',
        16: 'Core generative baseline for augmentation/synthesis pipelines.'
    },
    'project proposal and workplan.docx (6)': {
        9: 'Proposal/workplan document: defines expected outputs and milestones rather than final benchmark metrics.',
        12: 'Roadmap-based methodology, staged experiments, and validation planning.',
        14: 'Not a finalized experimental paper; outcomes depend on execution of plan.',
        15: 'N/A (planning document).',
        16: 'Useful for aligning scope, objectives, and evaluation path.'
    },
    'sensors-24-00641': {
        9: 'Reports improved physical-layer authentication performance using GAN-based augmentation.',
        12: 'Synthetic sample generation to rebalance/expand training distribution.',
        14: 'Synthetic-data mismatch may hurt robustness if generated distribution is biased.',
        15: 'Extra training-time overhead from GAN generation; inference usually unaffected.',
        16: 'Practical reference for augmentation in wireless security tasks.'
    },
    'sensors-25-07463': {
        9: 'Demonstrates Transformer + Soft Actor-Critic for UAV path optimization under precision-agriculture IoT constraints.',
        12: 'RL policy optimization with attention-based state representation.',
        14: 'RL approaches can be sample-inefficient and sensitive to reward/design choices.',
        15: 'Training can be long; online policy inference is typically faster.',
        16: 'Good reference if you extend CKM ideas toward control/planning.'
    },
    'SRC_SIg_21': {
        9: 'Shows GAN-based augmentation benefits for deep-learning localization systems.',
        12: 'Data synthesis to mitigate limited labeled localization data.',
        14: 'Quality of synthetic data is critical; poor GAN convergence can reduce gains.',
        15: 'Added generation/training cost; deployed inference remains manageable.',
        16: 'Directly relevant for dataset expansion strategies.'
    },
    'Thesis Research_ Channel Knowledge Map Prediction': {
        9: 'Thesis-focused CKM prediction narrative with expected methodological and evaluation outcomes.',
        12: 'Combines CKM prediction objectives with UAV/6G scenario framing.',
        14: 'As thesis material, maturity may vary across chapters/experiments.',
        15: 'Not consistently reported as a single benchmark runtime.',
        16: 'High strategic relevance to your project direction.'
    },
    'UAV_height_distribution (1) (1)': {
        9: 'Analyzes UAV transmitter height distribution effects on communication behavior/performance.',
        12: 'Parameter/sensitivity analysis centered on altitude-related propagation effects.',
        14: 'May require careful re-calibration for different terrains and mobility regimes.',
        15: 'Typically lightweight analytical/simulation workload.',
        16: 'Useful supporting reference for UAV scenario assumptions.'
    }
}

for r in range(2, ws.max_row + 1):
    file_name = ws.cell(r, 1).value
    if file_name in manual:
        for col, value in manual[file_name].items():
            ws.cell(r, col, value)

wb.save(excel_path)
print('Manual enrichment applied to rows:', len(manual))
print('Workbook saved:', excel_path)
