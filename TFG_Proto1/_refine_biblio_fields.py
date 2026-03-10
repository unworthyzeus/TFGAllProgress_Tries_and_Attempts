import openpyxl

p = r'c:/TFG/TFG_Proto1/docs/excel/Comparison Table (SOA).xlsx'
wb = openpyxl.load_workbook(p)
ws = wb[wb.sheetnames[0]]

updates = {
    2: ('Pattern Recognition journal page lists Impact Factor 7.6', None),
    3: ('No official JIF (MICCAI / LNCS conference proceedings)', 'OpenAlex h-index: Olaf Ronneberger=52'),
    4: ('IEEE Communications Surveys & Tutorials journal (official JIF not exposed on fetched page)', 'OpenAlex h-index: Yong Zeng=67'),
    5: ('No official JIF (preprint / arXiv)', 'OpenAlex h-index: Fabian Jaensch=3'),
    6: ('No official JIF (preprint / arXiv)', 'OpenAlex h-index: Zijian Wu=33'),
    7: ('No official JIF (preprint / arXiv)', 'OpenAlex h-index: Xiucheng Wang=17'),
    9: ('Electronics Letters journal (official JIF not exposed on fetched page)', 'OpenAlex h-index: Changpeng Zhou=9'),
    10: ('IEEE Transactions on Wireless Communications journal (official JIF not exposed on fetched page)', 'OpenAlex h-index: Xiaoli Xu=30'),
    11: ('No official JIF (CVPR conference paper)', 'OpenAlex h-index: Phillip Isola=46'),
    13: ('Sensors journal page lists Impact Factor 3.5', 'OpenAlex h-index: Lamia Alhoraibi=5'),
    14: ('Sensors journal page lists Impact Factor 3.5', 'OpenAlex h-index: G.-G. Ge=142'),
    15: ('No official JIF (ACM conference/proceedings paper)', 'OpenAlex h-index: Joseph Boulis=1'),
}

for row, (jif, h) in updates.items():
    ws.cell(row, 5).value = jif
    if h is not None:
        ws.cell(row, 6).value = h

for row in range(2, ws.max_row + 1):
    if ws.cell(row, 5).value is None or str(ws.cell(row, 5).value).strip() == '':
        ws.cell(row, 5).value = 'Official JIF not reliably verified from open sources (check JCR)'
    if ws.cell(row, 6).value is None or str(ws.cell(row, 6).value).strip() == '':
        ws.cell(row, 6).value = 'Author h-index not available (no reliable author match)'

wb.save(p)
print('refined workbook saved')
