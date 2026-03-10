import re
from pathlib import Path
import openpyxl

excel_path = Path('docs/excel/Comparison Table (SOA).xlsx')
md_root = Path('docs/markdown')

wb = openpyxl.load_workbook(excel_path)
ws = wb[wb.sheetnames[0]]

manual_meta = {
    '1-s2.0-S0031320321003897-main (1)': ('2021', 'Pattern Recognition (Elsevier)', 'N/A (journal metric lookup required)', 'N/A (author-level metric lookup required)'),
    '1505.04597v1': ('2015', 'MICCAI / arXiv preprint', 'N/A (conference/preprint)', 'N/A (author-level metric lookup required)'),
    '2309.07460v2 (1) (1)': ('2023', 'IEEE tutorial/survey (preprint source)', 'N/A (journal venue not explicitly fixed in extracted markdown)', 'N/A (author-level metric lookup required)'),
    '2402.00878v1 (2)': ('2024', 'Dataset/Preprint (radio map estimation)', 'N/A (preprint/dataset paper)', 'N/A (author-level metric lookup required)'),
    '2504.09849v1 (1) (1)': ('2025', 'Dataset/Preprint (CKMImageNet)', 'N/A (preprint/dataset paper)', 'N/A (author-level metric lookup required)'),
    '2507.12166v1': ('2025', 'Dataset/Preprint (RadioDiff-3D)', 'N/A (preprint/dataset paper)', 'N/A (author-level metric lookup required)'),
    '3.+18015+-+Aniru+Abudu+Nigeria+(311-332)': ('2025', 'International Journal of Research and Applied Technology', 'N/A (journal metric lookup required)', 'N/A (author-level metric lookup required)'),
    'Electronics Letters - 2025 - Zhou - Vision Channel Knowledge Map‐Aide_00ae6c299c': ('2025', 'Electronics Letters (Wiley)', 'N/A (journal metric lookup required)', 'N/A (author-level metric lookup required)'),
    'How_Much_Data_Is_Needed_for_Channel_Knowledge_Map_Construction': ('2024', 'IEEE Transactions on Wireless Communications', 'N/A (journal metric lookup required)', 'N/A (author-level metric lookup required)'),
    'Isola_Image-To-Image_Translation_With_CVPR_2017_paper': ('2017', 'CVPR (IEEE/CVF)', 'N/A (conference paper)', 'N/A (author-level metric lookup required)'),
    'project proposal and workplan.docx (6)': ('2026', 'Project proposal / internal document', 'N/A', 'N/A'),
    'sensors-24-00641': ('2024', 'Sensors (MDPI)', 'N/A (journal metric lookup required)', 'N/A (author-level metric lookup required)'),
    'sensors-25-07463': ('2025', 'Sensors (MDPI)', 'N/A (journal metric lookup required)', 'N/A (author-level metric lookup required)'),
    'SRC_SIg_21': ('2021', 'Conference/Proceedings source', 'N/A (conference/proceedings)', 'N/A (author-level metric lookup required)'),
    'Thesis Research_ Channel Knowledge Map Prediction': ('2026', 'Thesis/Research document', 'N/A', 'N/A'),
    'UAV_height_distribution (1) (1)': ('2024', 'Technical report / manuscript', 'N/A', 'N/A'),
}

for r in range(2, ws.max_row + 1):
    key = ws.cell(r, 1).value
    if key in manual_meta:
        year, pub, jif, hidx = manual_meta[key]
        ws.cell(r, 3, year)
        ws.cell(r, 4, pub)
        ws.cell(r, 5, jif)
        ws.cell(r, 6, hidx)
    else:
        # fallback to non-empty placeholders if any row was missed
        if not ws.cell(r, 3).value:
            ws.cell(r, 3, 'N/A')
        if not ws.cell(r, 4).value:
            ws.cell(r, 4, 'N/A')
        if not ws.cell(r, 5).value:
            ws.cell(r, 5, 'N/A (lookup required)')
        if not ws.cell(r, 6).value:
            ws.cell(r, 6, 'N/A (lookup required)')

wb.save(excel_path)
print('Bibliographic fields updated for rows:', ws.max_row - 1)
print('Workbook:', excel_path)
