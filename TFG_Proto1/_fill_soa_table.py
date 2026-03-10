import re
from pathlib import Path
import openpyxl

base = Path('docs/markdown')
excel_path = Path('docs/excel/Comparison Table (SOA).xlsx')

# Header mapping
headers = [
    'File', 'Approach', 'Year', 'Publisher/Journal/book', 'JIF (Journal Impact Factor)',
    'H-Index (by person)', 'Title', 'Metrics', 'Output (Results compared groundtruth)',
    'Dataset (type and number of samples) and augmentation', 'Models used', 'Heuristics',
    'Type of communication (ground, drones, SAR)', 'Disadvantages', 'How fast is the generation',
    'Other (Genral feeling)'
]

model_keywords = [
    'GAN','cGAN','pix2pix','U-Net','UNet','CNN','ResNet','Transformer','LSTM','KNN','Kriging',
    'XGBoost','Random Forest','SVM','MIMO','RIS','IRS','CKM','CGM','OTFS','DDAM','ray tracing','DNN'
]
metric_keywords = [
    'MSE','RMSE','MAE','PSNR','SSIM','IoU','Accuracy','Precision','Recall','F1','AUC','BER','throughput','latency'
]
comm_keywords = ['UAV','drone','aerial','satellite','ground','V2I','V2X','SAR','mmWave','6G','wireless']


def find_main_md(folder: Path):
    mds = sorted(folder.glob('*.md'))
    if mds:
        return mds[0]
    nested = sorted(folder.rglob('*.md'))
    return nested[0] if nested else None


def extract_title(lines, fallback):
    for ln in lines[:80]:
        t = ln.strip()
        if t.startswith('#'):
            t2 = re.sub(r'^#+\s*', '', t).strip()
            if t2 and len(t2) > 8:
                return t2[:300]
    return fallback


def extract_year(text, fallback_name):
    years = re.findall(r'\b(19\d{2}|20\d{2})\b', text[:12000])
    if years:
        # prioritize recent and plausible publication years
        ys = [int(y) for y in years if 1990 <= int(y) <= 2026]
        if ys:
            return str(sorted(ys)[0])
    yn = re.findall(r'\b(19\d{2}|20\d{2})\b', fallback_name)
    return yn[0] if yn else ''


def extract_approach(text):
    low = text.lower()
    if 'channel knowledge map' in low or 'ckm' in low:
        return 'CKM/Environment-aware communications'
    if 'conditional adversarial' in low or 'pix2pix' in low or 'cgan' in low:
        return 'Conditional GAN image-to-image'
    if 'mmwave' in low and ('beam' in low or 'beamforming' in low):
        return 'mmWave beam/CKM-aided approach'
    if 'survey' in low or 'tutorial' in low:
        return 'Tutorial/Survey'
    return 'Paper-specific method (see title/abstract)'


def extract_models(text):
    found = []
    for k in model_keywords:
        if re.search(r'(?i)\b' + re.escape(k) + r'\b', text):
            found.append(k)
    return ', '.join(dict.fromkeys(found))[:400]


def extract_metrics(text):
    found = []
    for k in metric_keywords:
        if re.search(r'(?i)\b' + re.escape(k) + r'\b', text):
            found.append(k)
    return ', '.join(dict.fromkeys(found))[:300]


def extract_dataset(text):
    # heuristic sentence grab around dataset keywords
    m = re.search(r'(?is)(dataset|data set|data|measurements?|samples?).{0,220}', text)
    if m:
        s = re.sub(r'\s+', ' ', m.group(0)).strip()
        return s[:350]
    return ''


def infer_comm_type(text):
    low = text.lower()
    tags = []
    if any(k in low for k in ['uav','drone','aerial']): tags.append('drones')
    if 'sar' in low: tags.append('SAR')
    if any(k in low for k in ['ground','cell','wireless','v2i','v2x','mmwave','6g']): tags.append('ground')
    return ', '.join(dict.fromkeys(tags))


def extract_disadvantages(text):
    m = re.search(r'(?is)(limitation|limitations|drawback|drawbacks|challenge|challenges|open problem|future work).{0,260}', text)
    if m:
        return re.sub(r'\s+', ' ', m.group(0)).strip()[:380]
    return ''


def short_other(text):
    if 'tutorial' in text.lower() or 'survey' in text.lower():
        return 'Strong background/reference paper.'
    if 'abstract' in text.lower():
        return 'Structured paper with clear abstract and methodology.'
    return ''

folders = [p for p in sorted(base.iterdir()) if p.is_dir()]
rows = []
for folder in folders:
    md = find_main_md(folder)
    if not md:
        continue
    try:
        text = md.read_text(encoding='utf-8', errors='ignore')
    except Exception:
        continue
    lines = text.splitlines()

    file_name = folder.name
    title = extract_title(lines, file_name)
    year = extract_year(text, file_name)
    approach = extract_approach(text)
    models = extract_models(text)
    metrics = extract_metrics(text)
    dataset = extract_dataset(text)
    comm_type = infer_comm_type(text)
    disadv = extract_disadvantages(text)

    # publisher/journal heuristic from text
    pub = ''
    for key in ['IEEE', 'Elsevier', 'Springer', 'CVPR', 'Sensors', 'Electronics Letters', 'Nature']:
        if re.search(r'(?i)\b' + re.escape(key) + r'\b', text[:10000]):
            pub = key
            break

    rows.append([
        file_name,
        approach,
        year,
        pub,
        '',
        '',
        title,
        metrics,
        '',
        dataset,
        models,
        '',
        comm_type,
        disadv,
        '',
        short_other(text)
    ])

wb = openpyxl.load_workbook(excel_path)
ws = wb[wb.sheetnames[0]]

# clear previous body rows
if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)

for i, row in enumerate(rows, start=2):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

wb.save(excel_path)
print(f'Filled rows: {len(rows)}')
print(f'Workbook: {excel_path}')
